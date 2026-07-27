import json
import csv
import tempfile
import os
import zipfile
from openpyxl import load_workbook
from ioc_rejudge.export import (
    export_jsonl,
    export_csv,
    export_excel,
    _CONCLUSION_ORDER,
    _OUTPUT_FIELDS,
    _REVIEW_ORDER,
)
from ioc_rejudge.models import Conclusion

_EXPECTED_REVIEW_HEADERS = [
    "空", "ioc", "有效/误报", "判定原因", "评审建议", "告警信息",
    "submitter", "comment", "context", "md5", "md5_list", "路由", "处置",
    "作用范围", "保留URL", "Provider状态", "缺失必要来源",
]

_EXTENDED_FIELDS = [
    "route",
    "disposition",
    "scope_actions",
    "retained_urls",
    "provider_statuses",
    "evidence_origins",
    "missing_required_providers",
]


def _make_verdict(ioc="test.com"):
    return {
        "original_ioc": ioc,
        "ioc": ioc,
        "ioc_type": "domain",
        "ports": "",
        "record_count": 1,
        "conclusion": Conclusion.ALIVE_VALID.value,
        "malicious_nature": "direct",
        "activity_status": "recent",
        "confidence": "high",
        "review_suggestion": "skip",
        "candidate_label": "",
        "latest_material_activity_time": "2026-03-20 17:10:41",
        "latest_intel_update_time": "",
        "source_set": "sample-base",
        "family": "SilverFox",
        "tag": "",
        "evidence_a_detail": "context/comment [strong,direct]: test",
        "evidence_b_detail": "hash.time [normal]: test",
        "evidence_c_detail": "",
        "evidence_d_detail": "",
        "evidence_e_detail": "",
        "evidence_f_detail": "",
        "hit_evidence": "A=context; B=hash.time",
        "forbidden_labels": "none",
        "reason": "alive valid",
        "alert_info": "alert_name:APT Alert\nadd_date:2026-05-01\nupdate_date:2026-05-02\ncampaign:CampaignX\nmalicious_family:SilverFox\nmalicious_type:TROJAN",
        "submitter": "alice",
        "comment": "sample comment",
        "context": "sample context",
        "md5": "abc",
        "md5_list": "abc,def",
        "route": "standard",
        "disposition": "block",
        "scope_actions": [{"scope": "domain", "action": "block", "ioc": ioc}],
        "retained_urls": [],
        "provider_statuses": {"whois": "success", "ioc_info": "no_data"},
        "evidence_origins": [{"provider": "whois", "kind": "whois"}],
        "missing_required_providers": [],
    }


def test_export_jsonl():
    verdicts = [_make_verdict("evil.com"), _make_verdict("bad.com")]
    with tempfile.NamedTemporaryFile(mode="w", suffix=".jsonl", delete=False) as f:
        export_jsonl(verdicts, f.name)
        f.flush()
        with open(f.name, encoding="utf-8") as rf:
            lines = rf.readlines()
        assert len(lines) == 2
        data = json.loads(lines[0])
        assert data["ioc"] == "evil.com"
        assert data["conclusion"] == Conclusion.ALIVE_VALID.value
        assert isinstance(data["scope_actions"], list)
        assert isinstance(data["provider_statuses"], dict)
    os.unlink(f.name)


def test_export_csv():
    verdicts = [_make_verdict("evil.com")]
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        export_csv(verdicts, f.name)
        f.flush()
        with open(f.name, encoding="utf-8") as rf:
            reader = csv.DictReader(rf)
            rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["ioc"] == "evil.com"
        assert rows[0]["provider_statuses"] == json.dumps(
            {"ioc_info": "no_data", "whois": "success"},
            ensure_ascii=False,
            sort_keys=True,
        )
        assert json.loads(rows[0]["scope_actions"])[0]["action"] == "block"
    os.unlink(f.name)


def test_export_csv_sanitizes_illegal_control_chars():
    verdict = _make_verdict("evil.com")
    verdict["reason"] = "bad\x08reason"
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        export_csv([verdict], f.name)
        f.flush()
        with open(f.name, encoding="utf-8") as rf:
            reader = csv.DictReader(rf)
            rows = list(reader)
        assert rows[0]["reason"] == "badreason"
    os.unlink(f.name)


def test_export_fields_keep_legacy_order_before_new_fields():
    legacy_fields = [
        "ioc",
        "conclusion",
        "malicious_nature",
        "activity_status",
        "confidence",
        "review_suggestion",
        "candidate_label",
        "hit_evidence",
        "forbidden_labels",
        "reason",
    ]
    assert _OUTPUT_FIELDS[:len(legacy_fields)] == legacy_fields
    assert "original_ioc" in _OUTPUT_FIELDS[len(legacy_fields):]
    assert "evidence_a_detail" in _OUTPUT_FIELDS[len(legacy_fields):]
    assert _OUTPUT_FIELDS[-len(_EXTENDED_FIELDS):] == _EXTENDED_FIELDS


def test_export_empty():
    with tempfile.NamedTemporaryFile(mode="w", suffix=".jsonl", delete=False) as f:
        export_jsonl([], f.name)
        f.flush()
        with open(f.name) as rf:
            assert rf.read() == ""
    os.unlink(f.name)


def test_export_excel():
    verdicts = [_make_verdict("evil.com"), _make_verdict("bad.com")]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel(verdicts, f.name)
        wb = load_workbook(f.name)
        assert wb.sheetnames == ["统计", "总", "判黑", "灰", "误报", "待复核"]
        ws = wb["总"]
        assert ws.max_row == 3  # header + 2 data rows
        expected = _EXPECTED_REVIEW_HEADERS
        assert [ws.cell(1, i).value for i in range(1, len(expected) + 1)] == expected
        # All three review sheets share the same header layout
        for sheet_name in ["判黑", "灰", "误报", "待复核"]:
            ws2 = wb[sheet_name]
            assert [ws2.cell(1, i).value for i in range(1, len(expected) + 1)] == expected
        assert {ws.cell(2, 2).value, ws.cell(3, 2).value} == {"evil.com", "bad.com"}
        assert ws.cell(2, 1).value in (None, "")
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None
        wb.close()
    os.unlink(f.name)


def test_export_excel_sanitizes_illegal_control_chars():
    verdict = _make_verdict("bad-char.com")
    verdict["alert_info"] = "alert\x00name"
    verdict["comment"] = "bad\x08comment"
    verdict["context"] = "bad\x0bcontext"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel([verdict], f.name)
        wb = load_workbook(f.name)
        ws = wb["总"]
        headers = [ws.cell(1, i).value for i in range(1, len(_EXPECTED_REVIEW_HEADERS) + 1)]
        comment_col = headers.index("comment") + 1
        context_col = headers.index("context") + 1
        alert_col = headers.index("告警信息") + 1
        assert ws.cell(2, comment_col).value == "badcomment"
        assert ws.cell(2, context_col).value == "badcontext"
        assert ws.cell(2, alert_col).value == "alertname"
        wb.close()
    os.unlink(f.name)


def test_export_excel_sanitizes_xml_unsafe_chars_in_sheet_xml():
    verdict = _make_verdict("xml-bad.com")
    verdict["comment"] = "c1\x85control"
    verdict["context"] = "surrogate\ud800value"
    verdict["alert_info"] = "nonchar\ufffevalue"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel([verdict], f.name)
        with zipfile.ZipFile(f.name) as zf:
            worksheet_xml = [
                zf.read(name).decode("utf-8")
                for name in zf.namelist()
                if name.startswith("xl/worksheets/sheet")
            ]
        combined = "\n".join(worksheet_xml)
        assert "\x85" not in combined
        assert "\ud800" not in combined
        assert "\ufffe" not in combined
        wb = load_workbook(f.name)
        ws = wb["总"]
        headers = [ws.cell(1, i).value for i in range(1, len(_EXPECTED_REVIEW_HEADERS) + 1)]
        comment_col = headers.index("comment") + 1
        context_col = headers.index("context") + 1
        alert_col = headers.index("告警信息") + 1
        assert ws.cell(2, comment_col).value == "c1control"
        assert ws.cell(2, context_col).value == "surrogatevalue"
        assert ws.cell(2, alert_col).value == "noncharvalue"
        wb.close()
    os.unlink(f.name)


def test_export_excel_empty():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel([], f.name)
        wb = load_workbook(f.name)
        assert wb.sheetnames == ["统计", "总", "判黑", "灰", "误报", "待复核"]
        ws = wb["总"]
        assert ws.max_row == 1  # header only
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is None  # no filter when no data
        wb.close()
    os.unlink(f.name)


def test_export_excel_review_sort():
    v1 = _make_verdict("a.com")
    v1["review_suggestion"] = list(_REVIEW_ORDER)[2]
    v1["conclusion"] = list(_CONCLUSION_ORDER)[3]
    v2 = _make_verdict("b.com")
    v2["review_suggestion"] = list(_REVIEW_ORDER)[0]
    v2["conclusion"] = list(_CONCLUSION_ORDER)[0]
    v3 = _make_verdict("c.com")
    v3["review_suggestion"] = list(_REVIEW_ORDER)[1]
    v3["conclusion"] = list(_CONCLUSION_ORDER)[1]
    verdicts = [v1, v2, v3]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel(verdicts, f.name)
        wb = load_workbook(f.name)
        ws = wb["总"]
        # Should be sorted: 蹇呯湅 -> 鎶芥 -> 涓嶇湅
        assert ws.cell(2, 2).value == "b.com"  # 蹇呯湅
        assert ws.cell(3, 2).value == "c.com"  # 鎶芥
        assert ws.cell(4, 2).value == "a.com"  # 涓嶇湅
        wb.close()
    os.unlink(f.name)


def test_export_excel_keeps_similar_indicators_adjacent():
    v1 = _make_verdict("evil.com")
    v1["review_suggestion"] = list(_REVIEW_ORDER)[1]
    v1["conclusion"] = list(_CONCLUSION_ORDER)[1]
    v2 = _make_verdict("a-other.com")
    v2["review_suggestion"] = list(_REVIEW_ORDER)[1]
    v2["conclusion"] = list(_CONCLUSION_ORDER)[1]
    v3 = _make_verdict("http://evil.com/path")
    v3["ioc"] = "evil.com/path"
    v3["ioc_type"] = "url"
    v3["review_suggestion"] = list(_REVIEW_ORDER)[1]
    v3["conclusion"] = list(_CONCLUSION_ORDER)[1]
    verdicts = [v1, v2, v3]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel(verdicts, f.name)
        wb = load_workbook(f.name)
        ws = wb["总"]
        ioc_col = _EXPECTED_REVIEW_HEADERS.index("ioc") + 1
        exported = [ws.cell(row, ioc_col).value for row in range(2, 5)]
        assert exported.index("evil.com") + 1 == exported.index("evil.com/path")
        wb.close()
    os.unlink(f.name)


def test_export_excel_with_diagnostics():
    verdicts = [_make_verdict("evil.com")]
    diag = {"processed_count": 1, "parse_error_count": 0, "skipped_total": 0}
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel(verdicts, f.name, diagnostics=diag)
        wb = load_workbook(f.name)
        ws = wb["统计"]
        # Summary should have data
        assert ws.max_row >= 2
        wb.close()
    os.unlink(f.name)


def test_export_excel_splits_black_and_false_positive_sheets():
    alive = _make_verdict("alive.com")
    alive["conclusion"] = "存活有效"
    inactive = _make_verdict("inactive.com")
    inactive["conclusion"] = "失活有效"
    pending = _make_verdict("pending.com")
    pending["conclusion"] = "待复核"
    pending["disposition"] = "review"
    gray = _make_verdict("gray.com")
    gray["conclusion"] = "灰"
    gray["disposition"] = "gray"
    gray["retained_urls"] = ["https://gray.com/phish"]
    false_positive = _make_verdict("fp.com")
    false_positive["conclusion"] = "误报"
    false_positive["disposition"] = "false_positive"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel([alive, inactive, pending, gray, false_positive], f.name)
        wb = load_workbook(f.name)
        black = wb["判黑"]
        gray_sheet = wb["灰"]
        false_sheet = wb["误报"]
        review_sheet = wb["待复核"]
        black_iocs = {black.cell(row, 2).value for row in range(2, black.max_row + 1)}
        gray_iocs = {gray_sheet.cell(row, 2).value for row in range(2, gray_sheet.max_row + 1)}
        fp_iocs = {false_sheet.cell(row, 2).value for row in range(2, false_sheet.max_row + 1)}
        review_iocs = {review_sheet.cell(row, 2).value for row in range(2, review_sheet.max_row + 1)}
        assert black_iocs == {"alive.com", "inactive.com"}
        assert gray_iocs == {"gray.com"}
        assert fp_iocs == {"fp.com"}
        assert review_iocs == {"pending.com"}
        wb.close()
    os.unlink(f.name)


def test_export_excel_structured_fields_and_route_provider_stats():
    verdict = _make_verdict("stats.invalid")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel([verdict], f.name)
        wb = load_workbook(f.name)
        total = wb["总"]
        headers = [cell.value for cell in total[1]]
        scope_value = total.cell(2, headers.index("作用范围") + 1).value
        provider_value = total.cell(2, headers.index("Provider状态") + 1).value
        assert json.loads(scope_value) == verdict["scope_actions"]
        assert json.loads(provider_value) == verdict["provider_statuses"]

        stats_values = [
            (wb["统计"].cell(row, 1).value, wb["统计"].cell(row, 2).value)
            for row in range(1, wb["统计"].max_row + 1)
        ]
        assert ("  standard", 1) in stats_values
        assert ("  ioc_info:no_data", 1) in stats_values
        assert ("  whois:success", 1) in stats_values
        wb.close()
    os.unlink(f.name)


def test_export_neutralizes_spreadsheet_formula_prefixes():
    payload = '=HYPERLINK("http://attacker.invalid","x")'
    verdict = _make_verdict("formula.invalid")
    verdict["family"] = payload
    verdict["comment"] = "+cmd|calc"
    verdict["context"] = "-2+3+cmd"
    verdict["submitter"] = "@attacker"
    verdict["md5"] = "\tTAB"

    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
        export_csv([verdict], f.name)
    with open(f.name, encoding="utf-8", newline="") as rf:
        rows = list(csv.reader(rf))
    os.unlink(f.name)
    header, data = rows[0], rows[1]
    assert data[header.index("family")] == "'" + payload
    assert data[header.index("ioc")] == "formula.invalid"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel([verdict], f.name)
        wb = load_workbook(f.name)
        ws = wb["总"]
        headers = [cell.value for cell in ws[1]]
        comment_cell = ws.cell(2, headers.index("comment") + 1)
        assert comment_cell.value == "'+cmd|calc"
        assert comment_cell.data_type == "s"
        assert ws.cell(2, headers.index("context") + 1).value == "'-2+3+cmd"
        assert ws.cell(2, headers.index("submitter") + 1).value == "'@attacker"
        assert ws.cell(2, headers.index("md5") + 1).value == "'\tTAB"
        wb.close()
    os.unlink(f.name)


def test_export_jsonl_preserves_raw_values_without_neutralization():
    payload = '=HYPERLINK("http://attacker.invalid","x")'
    verdict = _make_verdict("formula.invalid")
    verdict["family"] = payload

    with tempfile.NamedTemporaryFile(suffix=".jsonl", delete=False) as f:
        export_jsonl([verdict], f.name)
    with open(f.name, encoding="utf-8") as rf:
        row = json.loads(rf.readline())
    os.unlink(f.name)
    assert row["family"] == payload


def test_excel_review_sheets_include_reason_and_review_columns():
    verdict = _make_verdict("pending.invalid")
    verdict["conclusion"] = "待复核"
    verdict["reason"] = "必要样本查询未完整"
    verdict["review_suggestion"] = "必看"
    verdict["missing_required_providers"] = ["ioc_info", "fdark"]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        export_excel([verdict], f.name)
        wb = load_workbook(f.name)
        for sheet_name in ("总", "待复核"):
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            assert headers[2:5] == ["有效/误报", "判定原因", "评审建议"]
            assert headers[-1] == "缺失必要来源"
            assert ws.cell(2, headers.index("判定原因") + 1).value == "必要样本查询未完整"
            assert ws.cell(2, headers.index("评审建议") + 1).value == "必看"
            missing_value = ws.cell(2, headers.index("缺失必要来源") + 1).value
            assert json.loads(missing_value) == ["ioc_info", "fdark"]
        wb.close()
    os.unlink(f.name)
