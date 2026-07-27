"""Export verdicts to JSONL, CSV, and Excel formats."""
import json
import csv
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


_OUTPUT_FIELDS = [
    "ioc",
    "conclusion",
    "malicious_nature",
    "activity_status",
    "confidence",
    "review_suggestion",
    "candidate_label",
    "hit_evidence",
    "forbidden_labels",
    "reason",
    "original_ioc",
    "ioc_type",
    "ports",
    "record_count",
    "latest_material_activity_time",
    "latest_intel_update_time",
    "source_set",
    "family",
    "tag",
    "evidence_a_detail",
    "evidence_b_detail",
    "evidence_c_detail",
    "evidence_d_detail",
    "evidence_e_detail",
    "evidence_f_detail",
    "profile_observation_detail",
    "profile_domain_summary",
    "profile_ip_summary",
    "profile_runtime_summary",
    "threat_residue",
    "threat_residue_detail",
    "route",
    "disposition",
    "scope_actions",
    "retained_urls",
    "provider_statuses",
    "evidence_origins",
    "missing_required_providers",
]

_STRUCTURED_DEFAULTS = {
    "scope_actions": [],
    "retained_urls": [],
    "provider_statuses": {},
    "evidence_origins": [],
    "missing_required_providers": [],
}

_EXCEL_REVIEW_FIELDS = [
    "空",
    "ioc",
    "有效/误报",
    "判定原因",
    "评审建议",
    "告警信息",
    "submitter",
    "comment",
    "context",
    "md5",
    "md5_list",
    "路由",
    "处置",
    "作用范围",
    "保留URL",
    "Provider状态",
    "缺失必要来源",
]

_EXCEL_REVIEW_KEYS = [
    "review_blank",
    "ioc",
    "conclusion",
    "reason",
    "review_suggestion",
    "alert_info",
    "submitter",
    "comment",
    "context",
    "md5",
    "md5_list",
    "route",
    "disposition",
    "scope_actions",
    "retained_urls",
    "provider_statuses",
    "missing_required_providers",
]

_REVIEW_ORDER = {"必看": 0, "抽检": 1, "不看": 2}
_CONCLUSION_ORDER = {"待复核": 0, "灰": 1, "存活有效": 2, "失活有效": 3, "误报": 4}
_BLACK_CONCLUSIONS = {"存活有效", "失活有效"}

_FILL_REVIEW_REQUIRED = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
_FILL_CONCLUSION = {
    "存活有效": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "失活有效": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "灰": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    "误报": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "待复核": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

def export_jsonl(verdicts: list[dict], filepath: str):
    with open(filepath, "w", encoding="utf-8") as f:
        for v in verdicts:
            row = {field: _jsonl_value(v, field) for field in _OUTPUT_FIELDS}
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def export_csv(verdicts: list[dict], filepath: str):
    with open(filepath, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_OUTPUT_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for v in verdicts:
            writer.writerow({
                field: _neutralize_formula(_safe_value(v, field))
                for field in _OUTPUT_FIELDS
            })


def _raw_value(v, key: str, default=None):
    if isinstance(v, dict):
        return v.get(key, default)
    return getattr(v, key, default)


def _sanitize_json_value(value):
    if isinstance(value, str):
        return _sanitize_cell_text(value)
    if isinstance(value, list):
        return [_sanitize_json_value(item) for item in value]
    if isinstance(value, tuple):
        return [_sanitize_json_value(item) for item in value]
    if isinstance(value, dict):
        return {
            _sanitize_cell_text(str(key)): _sanitize_json_value(item)
            for key, item in value.items()
        }
    return value


def _jsonl_value(v, key: str):
    if key in _STRUCTURED_DEFAULTS:
        value = _raw_value(v, key, _STRUCTURED_DEFAULTS[key])
        if value is None:
            value = _STRUCTURED_DEFAULTS[key]
        return _sanitize_json_value(value)
    return _safe_value(v, key)


def _safe_value(v, key: str, default: str = ""):
    """Extract a string value from a verdict, whether it's a dict or dataclass."""
    structured = key in _STRUCTURED_DEFAULTS
    fallback = _STRUCTURED_DEFAULTS[key] if structured else default
    val = _raw_value(v, key, fallback)
    if val is None:
        val = fallback
    if structured and isinstance(val, (dict, list, tuple)):
        sanitized = _sanitize_json_value(val)
        return json.dumps(sanitized, ensure_ascii=False, sort_keys=True)
    return _sanitize_cell_text(str(val))


def _sanitize_cell_text(text: str) -> str:
    """Remove characters that are unsafe in Excel's worksheet XML."""
    return "".join(ch for ch in text if _is_excel_xml_char(ch))


def _is_excel_xml_char(ch: str) -> bool:
    code = ord(ch)
    if code in (0x09, 0x0A, 0x0D):
        return True
    if code < 0x20:
        return False
    if 0x7F <= code <= 0x9F:
        return False
    if 0xD800 <= code <= 0xDFFF:
        return False
    if code in (0xFFFE, 0xFFFF):
        return False
    return code <= 0x10FFFF


_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _neutralize_formula(text: str) -> str:
    """Prefix spreadsheet formula triggers so exported cells stay literal text."""
    if text and text.startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def _display_width(text: str) -> int:
    """Estimate display width: CJK chars count as 2, others as 1."""
    return sum(2 if "一" <= ch <= "鿿" or "　" <= ch <= "〿" or "＀" <= ch <= "￯" else 1 for ch in text)


def _sort_key(v: dict) -> tuple:
    """Sort key for review workflow: 必看 -> 抽检 -> 不看, then conclusion order."""
    review = _REVIEW_ORDER.get(_safe_value(v, "review_suggestion"), 9)
    conclusion = _CONCLUSION_ORDER.get(_safe_value(v, "conclusion"), 9)
    host_or_ip, port = _indicator_sort_parts(v)
    return (review, conclusion, host_or_ip, port, _safe_value(v, "ioc_type"), _safe_value(v, "ioc"))


def _indicator_sort_parts(v: dict) -> tuple[str, str]:
    ioc = _safe_value(v, "ioc")
    original_ioc = _safe_value(v, "original_ioc")
    raw = original_ioc or ioc

    if raw.lower().startswith(("http://", "https://")):
        parsed = urlparse(raw)
        host = (parsed.hostname or "").rstrip(".").lower()
        port = str(parsed.port or v.get("ports", "") or "")
        return host, port

    candidate = ioc.split("/", 1)[0]
    if ":" in candidate:
        host, port = candidate.rsplit(":", 1)
        if port.isdigit():
            return host.rstrip(".").lower(), port
    return candidate.rstrip(".").lower(), str(v.get("ports", "") or "")


def _auto_width(ws, fields: list[str]):
    """Set column widths based on content, CJK-aware."""
    for col_idx, field in enumerate(fields, 1):
        max_len = _display_width(field)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, _display_width(str(cell.value)))
        adjusted = min(max(max_len + 2, 8), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _add_stats_sheet(wb: Workbook, verdicts: list[dict], diagnostics: dict | None):
    """Add statistics sheet with counts and diagnostics."""
    ws = wb.create_sheet("统计", 0)
    ws.append(["IOC Rejudge Statistics"])
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=14)

    black_count = sum(1 for v in verdicts if v.get("conclusion", "") in _BLACK_CONCLUSIONS)
    false_positive_count = sum(1 for v in verdicts if v.get("conclusion", "") == "误报")

    ws.append(["Total Processed IOCs", len(verdicts)])
    ws.append(["判黑", black_count])
    ws.append(["灰", sum(1 for v in verdicts if v.get("conclusion", "") == "灰")])
    ws.append(["误报", false_positive_count])
    ws.append(["待复核", sum(1 for v in verdicts if v.get("conclusion", "") == "待复核")])

    # Counts by conclusion
    conclusion_counts: dict[str, int] = {}
    for v in verdicts:
        c = v.get("conclusion", "")
        conclusion_counts[c] = conclusion_counts.get(c, 0) + 1

    ws.append(["", ""])
    ws.append(["Conclusion Counts", ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    for conclusion, count in sorted(conclusion_counts.items(), key=lambda x: -x[1]):
        ws.append([f"  {conclusion}", count])

    # Counts by review suggestion
    review_counts: dict[str, int] = {}
    for v in verdicts:
        r = v.get("review_suggestion", "")
        review_counts[r] = review_counts.get(r, 0) + 1

    ws.append(["", ""])
    ws.append(["Review Suggestion Counts", ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    for suggestion, count in sorted(review_counts.items(), key=lambda x: _REVIEW_ORDER.get(x[0], 9)):
        ws.append([f"  {suggestion}", count])

    route_counts: dict[str, int] = {}
    provider_counts: dict[str, int] = {}
    for verdict in verdicts:
        route = str(verdict.get("route", "") or "")
        if route:
            route_counts[route] = route_counts.get(route, 0) + 1
        statuses = verdict.get("provider_statuses", {})
        if isinstance(statuses, str):
            try:
                statuses = json.loads(statuses)
            except json.JSONDecodeError:
                statuses = {}
        if isinstance(statuses, dict):
            for provider, status in statuses.items():
                key = f"{provider}:{status}"
                provider_counts[key] = provider_counts.get(key, 0) + 1

    ws.append(["", ""])
    ws.append(["Route Counts", ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    for route, count in sorted(route_counts.items()):
        ws.append([f"  {route}", count])

    ws.append(["", ""])
    ws.append(["Provider Status Counts", ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    for provider_status, count in sorted(provider_counts.items()):
        ws.append([f"  {provider_status}", count])

    # Diagnostics
    if diagnostics:
        ws.append(["", ""])
        ws.append(["Diagnostics", ""])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        diag_fields = [
            ("Parse Errors", "parse_error_count"),
            ("Missing Data", "missing_data_count"),
            ("Empty Data", "empty_data_count"),
            ("Non-list Data", "non_list_data_count"),
            ("No IOC", "no_ioc_count"),
            ("Total Skipped", "skipped_total"),
        ]
        for label, key in diag_fields:
            ws.append([f"  {label}", diagnostics.get(key, 0)])

        provider_metrics = diagnostics.get("provider_metrics", {})
        if isinstance(provider_metrics, dict) and provider_metrics:
            ws.append(["", ""])
            ws.append(["Provider Diagnostics", ""])
            ws[f"A{ws.max_row}"].font = Font(bold=True)
            for provider, metrics in sorted(provider_metrics.items()):
                if not isinstance(metrics, dict):
                    continue
                for metric, value in sorted(metrics.items()):
                    ws.append([f"  {provider}.{metric}", value])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


def _add_review_sheet(wb: Workbook, title: str, verdicts: list[dict]):
    """Add review-oriented sheet with the business column layout."""
    ws = wb.create_sheet(title)
    sorted_verdicts = sorted(verdicts, key=_sort_key)

    ws.append(_EXCEL_REVIEW_FIELDS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for v in sorted_verdicts:
        ws.append([
            _neutralize_formula(_safe_value(v, key))
            for key in _EXCEL_REVIEW_KEYS
        ])

    for row_idx in range(2, len(sorted_verdicts) + 2):
        conclusion_val = ws.cell(row=row_idx, column=_EXCEL_REVIEW_FIELDS.index("有效/误报") + 1).value
        if conclusion_val in _FILL_CONCLUSION:
            for cell in ws[row_idx]:
                cell.fill = _FILL_CONCLUSION[conclusion_val]

    _auto_width(ws, _EXCEL_REVIEW_FIELDS)
    ws.freeze_panes = "A2"

    if sorted_verdicts:
        last_col = get_column_letter(len(_EXCEL_REVIEW_FIELDS))
        ws.auto_filter.ref = f"A1:{last_col}{len(sorted_verdicts) + 1}"


def export_excel(
    verdicts: list[dict],
    filepath: str,
    diagnostics: dict | None = None,
):
    """Export Excel workbook with statistics and review sheets.

    Args:
        verdicts: List of verdict dicts.
        filepath: Output .xlsx path.
        diagnostics: Optional diagnostics dict for statistics sheet.
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    _add_stats_sheet(wb, verdicts, diagnostics)
    _add_review_sheet(wb, "总", verdicts)
    _add_review_sheet(
        wb,
        "判黑",
        [v for v in verdicts if _safe_value(v, "conclusion") in _BLACK_CONCLUSIONS],
    )
    _add_review_sheet(
        wb,
        "灰",
        [v for v in verdicts if _safe_value(v, "conclusion") == "灰"],
    )
    _add_review_sheet(
        wb,
        "误报",
        [v for v in verdicts if _safe_value(v, "conclusion") == "误报"],
    )
    _add_review_sheet(
        wb,
        "待复核",
        [v for v in verdicts if _safe_value(v, "conclusion") == "待复核"],
    )

    wb.save(filepath)
